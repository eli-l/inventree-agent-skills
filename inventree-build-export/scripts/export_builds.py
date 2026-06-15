#!/usr/bin/env python3
"""export_builds.py — Export InvenTree Build Orders to XLSX.

Schema: one row per (BO × consumed child StockItem).
The same BO can produce multiple consumed rows — one per StockItem
that the BO drained, e.g. a 4-side-panel BO has 4 child rows, a
filament print that pulled from 2 spools has 2 child rows.

Each row shows BOTH sides of the build:
  - PRODUCED (constant across all rows of one BO)
  - CONSUMED (varies per row)

References use IPN throughout. Where IPN is missing, the cell shows
"<name>@pk<n>" as a fallback so the row is still uniquely referenceable.

Output is XLSX, written to OUTFILE (default inventree_builds_<timestamp>.xlsx).

Usage:
  export_builds.py [OUTFILE.xlsx] [--url URL] [--token ***

INV_TOKEN and INV_URL may be set in the environment.

Requires: openpyxl (pip install openpyxl).
"""
import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def load_config(token=None, url=None):
    token = token or os.environ.get("INV_TOKEN")
    url = url or os.environ.get("INV_URL")
    if not token:
        sys.exit("INV_TOKEN missing. Pass --token *** set $INV_TOKEN.")
    if not url:
        sys.exit("INV_URL missing. Pass --url or set $INV_URL.")
    return token, url.rstrip("/")


class InvenTree:
    def __init__(self, token, url, page_size=200):
        self.token = token
        self.url = url.rstrip("/")
        self.page_size = page_size

    def _req(self, method, path, body=None):
        headers = {
            "Authorization": f"Token {self.token}",
            "Referer": self.url,
            "Accept": "application/json",
        }
        data = None
        if body is not None:
            data = json.dumps(body).encode()
            headers["Content-Type"] = "application/json"
        req = urllib.request.Request(
            f"{self.url}{path}", data=data, headers=headers, method=method
        )
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())

    def get(self, path):
        return self._req("GET", path)

    def get_all(self, path):
        out = []
        offset = 0
        while True:
            sep = "&" if "?" in path else "?"
            d = self.get(f"{path}{sep}limit={self.page_size}&offset={offset}")
            items = d if isinstance(d, list) else d.get("results", [])
            if not items:
                break
            out.extend(items)
            if isinstance(d, dict) and d.get("next") is None:
                break
            offset += len(items)
            if len(out) > 10000:
                break
        return out


# Columns in display order: (key, header, tooltip)
COLUMNS = [
    # Build Order (anchor)
    ("bo_reference",        "BO Reference",        "Full reference, e.g. BO-0024"),
    ("bo_status_text",      "BO Status",           "Pending / Production / Complete"),
    ("bo_start_date",       "BO Started",          ""),
    ("bo_completion_date",  "BO Completed On",     ""),

    # PRODUCED (constant across all rows of one BO)
    ("produced_part_ipn",   "Produced Part IPN",   "IPN of the assembly this BO builds"),
    ("produced_part_name",  "Produced Part Name",  "Human-readable name"),
    ("produced_qty",        "Produced Qty",        "Total units produced (= sum of output StockItem quantities)"),
    ("production_cost",     "Production Cost",     "Currency-aggregated cost of producing the BO output, computed recursively: filament = qty x purchase_price; sub-assemblies = sum of their children's costs"),
    ("production_currency", "Currency",            "Currency for Production Cost (from source PO line, or 'mixed')"),
    ("output_stockitem_pks","Output StockItems",   "pk(s) of the StockItem(s) this BO created"),

    # BOM Sub-Part (constant across rows of one build_line)
    ("bom_sub_part_ipn",    "BOM Sub-Part IPN",    "Generic — part listed in the BOM (placeholder)"),
    ("bom_sub_part_name",   "BOM Sub-Part Name",   "Generic — part listed in the BOM"),
    ("bom_qty_per_unit",    "BOM Qty / BO Unit",   "Quantity of this sub_part per 1 BO unit"),

    # CONSUMED (varies per row)
    ("consumed_part_ipn",   "Consumed Part IPN",   "IPN of the specific material that was actually used"),
    ("consumed_part_name",  "Consumed Part Name",  "Specific — real filament / part"),
    ("consumed_qty",        "Consumed Qty",        "Quantity drained from this child StockItem"),
    ("consumed_stockitem_pk","Consumed SI pk",     "The child StockItem that was consumed"),
    ("consumed_cost",       "Consumed Cost",       "This child StockItem's contribution to Production Cost"),
    ("consumed_mfg_cost",   "Consumed Mfg Cost",   "This child Part's mfg-cost contribution (consumed_qty × per-unit mfg cost)"),

    # SOURCE (varies per row)
    ("source_type",         "Source Type",         "purchase_order (filament) | build (sub-assembly) | manual | data_mismatch | (blank=not consumed)"),
    ("source_ref",          "Source Ref",          "PO ref for filament, BO ref for sub-assembly"),
    ("source_stockitem_pk", "Source SI pk",        "StockItem the child was drawn from"),
    ("source_batch",        "Source Batch",        ""),

    # TOTAL (per BO)
    ("mfg_cost",            "Mfg Cost",            "Manufacturing cost rollup for the produced part (own cost entries + sub-assembly mfg costs), multiplied by produced_qty. Source: /plugin/manufacturing-costs/"),
    ("total_cost",          "Total Cost",          "Production Cost + Mfg Cost"),
]


def part_ref(part):
    """IPN if available, else '<name>@pk<n>' as a stable fallback."""
    if not part:
        return ""
    ipn = (part.get("IPN") or "").strip()
    if ipn:
        return ipn
    name = part.get("name") or "?"
    return f"{name}@pk{part.get('pk', '?')}"


def part_name(part):
    return (part or {}).get("name", "") or ""


def export(inv, outfile):
    print("Fetching all Build Orders ...", file=sys.stderr)
    builds = inv.get_all("/api/build/")

    # Pre-fetch all parts (we'll look up by pk throughout)
    print("Fetching all Parts ...", file=sys.stderr)
    parts = {p["pk"]: p for p in inv.get_all("/api/part/")}

    # Pre-fetch all stockitems
    print("Fetching all StockItems ...", file=sys.stderr)
    all_stock = inv.get_all("/api/stock/")

    # Pre-fetch all locations
    print("Fetching all Locations ...", file=sys.stderr)
    locations = {l["pk"]: l for l in inv.get_all("/api/stock/location/")}

    # Pre-fetch all BOMs (for the manufacturing-cost sub-assembly rollup)
    print("Fetching all BOMs ...", file=sys.stderr)
    all_boms = inv.get_all("/api/bom/")
    bom_by_parent = {}  # parent_part_pk -> [{sub_part, quantity}, ...]
    for b in all_boms:
        bom_by_parent.setdefault(b.get("part"), []).append({
            "sub_part": b.get("sub_part"),
            "quantity": float(b.get("quantity") or 0),
        })

    # Pre-fetch ManufacturingCosts plugin data (rates + cost entries).
    # Endpoint lives under /plugin/, not /api/, in InvenTree 1.3+.
    print("Fetching ManufacturingCosts plugin data ...", file=sys.stderr)
    rate_by_pk = {}                    # rate_pk -> {price, price_currency, units, name}
    cost_entries_by_part = {}          # part_pk -> [cost_entry, ...]
    try:
        rates = inv.get_all("/plugin/manufacturing-costs/rate/")
        for r in rates:
            rate_by_pk[r["pk"]] = r
        costs = inv.get_all("/plugin/manufacturing-costs/cost/")
        for c in costs:
            if not c.get("active", True):
                continue
            cost_entries_by_part.setdefault(c.get("part"), []).append(c)
    except Exception as e:
        print(f"  [warn] ManufacturingCosts plugin not reachable: {e}", file=sys.stderr)
    print(f"  rates: {len(rate_by_pk)}  cost entries: {sum(len(v) for v in cost_entries_by_part.values())}", file=sys.stderr)

    # Manufacturing cost rollup (per part, memoised). Returns the cost of
    # producing 1 unit of `part_pk` in EUR, including sub-assembly mfg
    # costs. Cycles are guarded.
    _mfg_memo = {}
    _mfg_visiting = set()

    def part_mfg_cost(part_pk):
        if part_pk in _mfg_memo:
            return _mfg_memo[part_pk]
        if part_pk in _mfg_visiting:
            return 0.0
        _mfg_visiting.add(part_pk)
        total = 0.0
        # 1. Direct cost entries on this part
        for ce in cost_entries_by_part.get(part_pk, []):
            rate = rate_by_pk.get(ce.get("rate"), {})
            price = float(rate.get("price") or 0)
            qty = float(ce.get("quantity") or 0)
            total += qty * price
        # 2. Sub-assembly mfg costs (recursive, multiplied by BOM qty)
        for b in bom_by_parent.get(part_pk, []):
            sub_pk = b["sub_part"]
            sub_qty = b["quantity"]
            total += sub_qty * part_mfg_cost(sub_pk)
        _mfg_visiting.discard(part_pk)
        _mfg_memo[part_pk] = total
        return total

    # Partition stockitems.
    # An "output" of a BO is a StockItem that was PRODUCED by it
    # (s["build"] is set, s["quantity"] > 0). Note: we do NOT filter by
    # consumed_by — sub-assemblies often have consumed_by set immediately
    # because the sub-assembly is consumed by the next BO in the chain
    # (e.g. FRAME-TF produced by BO-0008 is consumed by BO-0001's FRAME).
    # The output qty is still the produced qty; the consumption is a
    # separate transaction.
    output_by_build = {}    # bo_pk -> [output stockitems]
    consumed_by_build = {}  # bo_pk -> [consumed child stockitems]
    for s in all_stock:
        if s.get("build") and s.get("quantity", 0) > 0:
            output_by_build.setdefault(s["build"], []).append(s)
        if s.get("consumed_by"):
            consumed_by_build.setdefault(s["consumed_by"], []).append(s)

    # Cache for source lookups
    stockitems_by_pk = {s["pk"]: s for s in all_stock}
    bo_refs = {b["pk"]: b.get("reference", "") for b in builds}

    def get_bo_ref(bo_pk):
        return bo_refs.get(bo_pk, "")

    # Compute per-StockItem costs (recursive, cached). Done once for the
    # whole catalog so per-BO production cost is just a sum.
    print("Computing production costs (recursive) ...", file=sys.stderr)
    cost_memo = _compute_costs(consumed_by_build, stockitems_by_pk.get)

    def bo_production_cost(bo_pk):
        children = consumed_by_build.get(bo_pk, [])
        if not children:
            return (0.0, "")
        total = 0.0
        currencies = set()
        for c in children:
            c_cost, c_cur = cost_memo.get(c["pk"], (0.0, ""))
            total += c_cost
            if c_cur:
                currencies.add(c_cur)
        cur = currencies.pop() if len(currencies) == 1 else (
            "mixed" if currencies else ""
        )
        return (total, cur)

    def bo_mfg_cost(bo):
        """Manufacturing cost for the produced output of this BO.
        = produced_qty × mfg_cost_per_unit(produced_part).
        All rates today are EUR; if that changes, currency aggregation
        follows the same pattern as bo_production_cost.
        """
        produced_pk = bo.get("part")
        if not produced_pk:
            return 0.0
        produced_qty = float(bo.get("quantity") or 1) or 1
        return produced_qty * part_mfg_cost(produced_pk)

    print(f"Processing {len(builds)} BOs ...", file=sys.stderr)
    rows = []
    for b in builds:
        bo_pk = b["pk"]
        produced = parts.get(b.get("part")) or {}

        # Output (produced) info — constant for this BO
        outputs = output_by_build.get(bo_pk, [])
        produced_qty = sum(float(s.get("quantity") or 0) for s in outputs)
        output_pks = [str(s.get("pk", "")) for s in outputs]

        # Consumed children (input)
        children = consumed_by_build.get(bo_pk, [])

        # build_lines
        try:
            bls = inv.get(f"/api/build/line/?build={bo_pk}")
            bls = bls if isinstance(bls, list) else bls.get("results", [])
        except urllib.error.HTTPError:
            bls = []

        # Pair build_lines to children (1 bl can map to N children)
        pairings = pair_bls_to_children(bls, children)

        if not bls:
            # Pending BO with no build_lines — show produced + no consumption
            prod_cost, prod_cur = bo_production_cost(bo_pk)
            mfg_cost = bo_mfg_cost(b)
            rows.append(_row(
                b=b, produced=produced, produced_qty=produced_qty,
                production_cost=prod_cost, production_currency=prod_cur,
                output_pks=output_pks, bl=None, bom_sub_part=None,
                child=None, mfg_cost=mfg_cost, get_bo_ref=get_bo_ref,
            ))
            continue

        prod_cost, prod_cur = bo_production_cost(bo_pk)
        mfg_cost = bo_mfg_cost(b)
        for bl in bls:
            bom_sub_part = parts.get(bl.get("part")) or {}
            bl_children = pairings.get(bl["pk"], [])

            if not bl_children:
                rows.append(_row(
                    b=b, produced=produced, produced_qty=produced_qty,
                    production_cost=prod_cost, production_currency=prod_cur,
                    output_pks=output_pks, bl=bl, bom_sub_part=bom_sub_part,
                    child=None, mfg_cost=mfg_cost, get_bo_ref=get_bo_ref,
                ))
                continue

            for child in bl_children:
                c_cost, _ = cost_memo.get(child["pk"], (0.0, ""))
                # Per-row mfg cost contribution = consumed_qty × per-unit mfg cost of consumed part
                consumed_part_pk = child.get("part")
                consumed_qty = float(child.get("quantity") or 0)
                consumed_mfg = (consumed_qty * part_mfg_cost(consumed_part_pk)
                                if consumed_part_pk else 0.0)
                rows.append(_row(
                    b=b, produced=produced, produced_qty=produced_qty,
                    production_cost=prod_cost, production_currency=prod_cur,
                    output_pks=output_pks, bl=bl, bom_sub_part=bom_sub_part,
                    child=child, consumed_cost=c_cost,
                    consumed_mfg_cost=consumed_mfg, mfg_cost=mfg_cost,
                    get_bo_ref=get_bo_ref,
                ))

    rows.sort(key=_sort_key)
    _write_xlsx(outfile, rows)
    return outfile, len(rows), len(builds)


def pair_bls_to_children(bls, children):
    """Pair each build_line with all consumed children that belong to it.

    Returns {bl_pk: [child, ...]}. A single bl can have many children
    (e.g. filament from 2 spools of the same IPN).

    Strategy:
      1. Exact match: child.part == bl.part (specific filament in BOM)
      2. Fallback for unmatched bls: assign remaining unused children
         to whichever bl needs the most material (by bom_qty desc).
         Handles generic-BOM (e.g. "PETG") vs specific stock ("PETG White").
    """
    children_by_part = {}
    for c in children:
        children_by_part.setdefault(c.get("part"), []).append(c)

    used = set()  # child pks already assigned
    pairings = {bl["pk"]: [] for bl in bls}

    # Pass 1: exact part match
    for bl in bls:
        bl_part = bl.get("part")
        for c in children_by_part.get(bl_part, []):
            if c["pk"] not in used:
                pairings[bl["pk"]].append(c)
                used.add(c["pk"])

    # Pass 2: fallback for bls with no exact matches
    unmatched_bls = [bl for bl in bls if not pairings[bl["pk"]]]
    unused = [c for c in children if c["pk"] not in used]
    if unmatched_bls and unused:
        if len(unmatched_bls) == 1:
            # Single unmatched bl (generic BOM vs specific stock):
            # all remaining children go to it.
            pairings[unmatched_bls[0]["pk"]] = list(unused)
        else:
            # Multiple unmatched bls: distribute greedily by quantity
            # (largest bl picks the largest unused child, recursively).
            unmatched_bls.sort(key=lambda bl: -float(bl.get("quantity") or 0))
            remaining = list(unused)
            for bl in unmatched_bls:
                if not remaining:
                    break
                child = max(remaining, key=lambda c: float(c.get("quantity") or 0))
                pairings[bl["pk"]].append(child)
                remaining = [c for c in remaining if c["pk"] != child["pk"]]

    return pairings


def _resolve_source(child):
    """Find where a consumed child StockItem's content came from.

    Two patterns:
      1. New workflow: child has parent=<source>; source carries
         build or purchase_order.
      2. Old workflow: child IS the source (no parent); child itself
         carries build or purchase_order.
    """
    # We rely on the child + parent for this; both are in all_stock.
    parent_pk = child.get("parent")
    parent = None
    if parent_pk:
        # Re-fetch on demand (or use the global stockitems list)
        parent = _get_stockitem(parent_pk)
    source = parent if parent else child
    source_pk = parent_pk if parent_pk else child.get("pk", "")

    if not source:
        return {"source_type": "manual", "source_ref": "",
                "source_stockitem_pk": "", "source_batch": ""}
    parent_bo_pk = source.get("build")
    po_pk = source.get("purchase_order")
    if parent_bo_pk:
        stype = "build"
        ref = _get_bo_ref(parent_bo_pk)
    elif po_pk:
        stype = "purchase_order"
        ref = source.get("purchase_order_reference") or ""
    else:
        stype = "manual"
        ref = ""
    return {
        "source_type": stype,
        "source_ref": ref,
        "source_stockitem_pk": source_pk or "",
        "source_batch": source.get("batch") or "",
    }


def _compute_costs(consumed_by_build, _si_cache_ref):
    """Compute the cost of every StockItem in the instance, recursively.

    Algorithm:
      - If a StockItem has `purchase_price > 0` and no producing BO,
        it is filament (or other purchased stock) and its cost is
        `quantity * purchase_price` (the StockItem already stores the
        per-unit price in the same unit as `quantity`).
      - If a StockItem has a producing BO (`build` set, no
        `purchase_price`), it is a sub-assembly. Its cost is the
        sum of costs of the StockItems consumed by that BO.
      - Otherwise (no price, no BO, no children) cost is 0.

    Returns: dict[si_pk, (cost: float, currency: str)]. Currency is
    inherited from the filament source; 'mixed' if multiple currencies
    are summed; '' if no cost could be computed.
    """
    cost_memo = {}      # si_pk -> (cost, currency)
    visiting = set()    # cycle detection

    def cost_of(si_pk):
        if si_pk in cost_memo:
            return cost_memo[si_pk]
        if si_pk in visiting:
            # Cycle — should not happen in well-formed data, but guard.
            return (0.0, "")
        visiting.add(si_pk)
        si = _si_cache_ref(si_pk)
        if not si:
            visiting.discard(si_pk)
            cost_memo[si_pk] = (0.0, "")
            return cost_memo[si_pk]

        producing_bo = si.get("build")
        purchase_price = float(si.get("purchase_price") or 0)

        if not producing_bo and purchase_price > 0:
            # Filament (or other purchased stock) — direct cost.
            qty = float(si.get("quantity") or 0)
            cur = si.get("purchase_price_currency") or ""
            visiting.discard(si_pk)
            cost_memo[si_pk] = (qty * purchase_price, cur)
            return cost_memo[si_pk]

        if producing_bo:
            # Sub-assembly — sum children's costs.
            children = consumed_by_build.get(producing_bo, [])
            total = 0.0
            currencies = set()
            for c in children:
                c_cost, c_cur = cost_of(c["pk"])
                total += c_cost
                if c_cur:
                    currencies.add(c_cur)
            cur = currencies.pop() if len(currencies) == 1 else (
                "mixed" if currencies else ""
            )
            visiting.discard(si_pk)
            cost_memo[si_pk] = (total, cur)
            return cost_memo[si_pk]

        # No price, no BO — unknown.
        visiting.discard(si_pk)
        cost_memo[si_pk] = (0.0, "")
        return cost_memo[si_pk]

    # Walk every consumed child (that's our entry point — we only need
    # to compute costs for SIs that show up as consumed in some BO).
    for children in consumed_by_build.values():
        for c in children:
            cost_of(c["pk"])

    return cost_memo


# Caches populated during export() and used by _resolve_source().
_stockitems_by_pk = {}
_bo_refs_cache = {}


def _get_stockitem(pk):
    return _stockitems_by_pk.get(pk)


def _get_bo_ref(bo_pk):
    return _bo_refs_cache.get(bo_pk, "")


def _row(b, produced, produced_qty, production_cost, production_currency,
         output_pks, bl=None, bom_sub_part=None, child=None,
         consumed_cost=None, consumed_mfg_cost=None, mfg_cost=None,
         get_bo_ref=None):
    s = None
    if child:
        s = _resolve_source(child)
    s = s or {}

    consumed_part_pk = child.get("part") if child else None
    consumed_part = _parts_cache.get(consumed_part_pk, {}) if consumed_part_pk else {}
    consumed_qty = float(child.get("quantity") or 0) if child else ""
    consumed_si_pk = child.get("pk", "") if child else ""

    prod_cost = float(production_cost or 0)
    mfg = float(mfg_cost or 0)
    total_cost = round(prod_cost + mfg, 4)

    return {
        "bo_reference":        b.get("reference", ""),
        "bo_status_text":      b.get("status_text", ""),
        "bo_start_date":       b.get("start_date", "") or "",
        "bo_completion_date":  b.get("completion_date", "") or "",

        "produced_part_ipn":   part_ref(produced),
        "produced_part_name":  part_name(produced),
        "produced_qty":        produced_qty,
        "production_cost":     round(prod_cost, 4),
        "production_currency": production_currency or "",
        "output_stockitem_pks":";".join(output_pks),

        "bom_sub_part_ipn":    part_ref(bom_sub_part) if bom_sub_part else "",
        "bom_sub_part_name":   part_name(bom_sub_part) if bom_sub_part else "",
        "bom_qty_per_unit":    float(bl.get("quantity") or 0) if bl else "",

        "consumed_part_ipn":   part_ref(consumed_part),
        "consumed_part_name":  part_name(consumed_part),
        "consumed_qty":        consumed_qty,
        "consumed_stockitem_pk": consumed_si_pk,
        "consumed_cost":       round(float(consumed_cost or 0), 4) if consumed_cost is not None else "",
        "consumed_mfg_cost":   round(float(consumed_mfg_cost or 0), 4) if consumed_mfg_cost is not None else "",

        "source_type":         s.get("source_type", ""),
        "source_ref":          s.get("source_ref", ""),
        "source_stockitem_pk": s.get("source_stockitem_pk", ""),
        "source_batch":        s.get("source_batch", ""),

        "mfg_cost":            round(mfg, 4),
        "total_cost":          total_cost,
    }


# Module-level parts cache; populated during export().
_parts_cache = {}


def _sort_key(r):
    """Sort by BO descending, then by BOM sub-part IPN, then by consumed IPN."""
    ref = r.get("bo_reference", "")
    bo_num = 0
    if ref and "-" in ref:
        try:
            bo_num = int(ref.split("-")[-1])
        except ValueError:
            pass
    return (-bo_num,
            r.get("bom_sub_part_ipn", "") or "",
            r.get("consumed_part_ipn", "") or "")


def _write_xlsx(outfile, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Builds"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    keys = [k for k, _, _ in COLUMNS]
    headers = [h for _, h, _ in COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    zebra = PatternFill("solid", fgColor="F8F9FA")
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            val = r.get(key, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = zebra

    widths = {
        "bo_reference": 11, "bo_status_text": 11,
        "bo_start_date": 12, "bo_completion_date": 14,
        "produced_part_ipn": 28, "produced_part_name": 30,
        "produced_qty": 11, "production_cost": 12, "production_currency": 9,
        "output_stockitem_pks": 16,
        "bom_sub_part_ipn": 26, "bom_sub_part_name": 30, "bom_qty_per_unit": 11,
        "consumed_part_ipn": 28, "consumed_part_name": 32,
        "consumed_qty": 11, "consumed_stockitem_pk": 11, "consumed_cost": 12,
        "consumed_mfg_cost": 14,
        "source_type": 17, "source_ref": 11,
        "source_stockitem_pk": 11, "source_batch": 12,
        "mfg_cost": 12, "total_cost": 12,
    }
    for col_idx, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 14)

    ws.freeze_panes = "E2"
    ws.row_dimensions[1].height = 36
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{len(rows) + 1}"

    wb.save(outfile)


# Output directory: every run writes into ./export/ (gitignored).
EXPORT_DIR = "export"


def _resolve_outfile(outfile_arg):
    """All output lands in ./export/. The arg is treated as a filename
    only — any path component is dropped — so accidental
    'foo/../bar.xlsx' cannot write outside the export dir.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)
    name = os.path.basename(outfile_arg) if outfile_arg else None
    if not name or not name.endswith(".xlsx"):
        name = f"inventree_builds_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return os.path.join(EXPORT_DIR, name)


def main():
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument(
        "outfile", nargs="?",
        default=None,
        help="Output filename inside ./export/ (default: inventree_builds_<timestamp>.xlsx). "
             "Path components are ignored — output always lands in ./export/.",
    )
    p.add_argument("--url", help="InvenTree base URL (default: $INV_URL)")
    p.add_argument("--token", help="InvenTree API token (default: $INV_TOKEN)")
    p.add_argument("--page-size", type=int, default=200, help="Pagination page size")
    args = p.parse_args()

    outfile = _resolve_outfile(args.outfile)

    token, url = load_config(args.token, args.url)
    inv = InvenTree(token, url, page_size=args.page_size)
    print(f"Connecting to {url} ...", file=sys.stderr)

    # Pre-load the global caches _resolve_source/_row will use
    global _stockitems_by_pk, _bo_refs_cache, _parts_cache
    print("Pre-loading stockitems for source resolution ...", file=sys.stderr)
    _stockitems_by_pk = {s["pk"]: s for s in inv.get_all("/api/stock/")}
    _bo_refs_cache = {b["pk"]: b.get("reference", "")
                      for b in inv.get_all("/api/build/")}
    print("Pre-loading parts ...", file=sys.stderr)
    _parts_cache = {p["pk"]: p for p in inv.get_all("/api/part/")}

    outfile, n_rows, n_builds = export(inv, outfile=outfile)
    print(f"Exported {n_rows} rows for {n_builds} Build Orders -> {outfile}", file=sys.stderr)


if __name__ == "__main__":
    main()
